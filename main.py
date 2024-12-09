import logging
import os
import random
import re
import threading
import time
import tkinter as tk
from datetime import datetime
from tkinter import messagebox, ttk
from typing import Dict, List, Tuple
from urllib.parse import parse_qs, unquote, urlparse
import csv
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter

# Constants for search operators and engines
SEARCH_OPERATORS: Dict[str, str] = {
    "site:": "Search for pages from a specific website",
    "inurl:": "Search for a term in the URL of a page",
    "intitle:": "Search for a term in the title of a page",
    "intext:": "Search for a term in the text of a page",
    "filetype:": "Search for a specific file type",
    "author:": "Search for content by a specific author",
    "source:": "Search for content from a specific source",
    "location:": "Search for content related to a specific location",
    "before:": "Search for content published before a specific date",
    "after:": "Search for content published after a specific date",
}

SEARCH_ENGINES: List[Tuple[str, str]] = [
    ("Bing", "scrape_bing"),
    ("DuckDuckGo", "scrape_duckduckgo"),
    ("Yahoo", "scrape_yahoo"),
    ("Mojeek", "scrape_mojeek"),  # TODO: Implement Mojeek scraper
]

class SearchScraperGUI:
    def __init__(self, master: tk.Tk):
        self.master = master
        master.title("Search Scraper")
        self.total_pages = 0
        self.scraped_pages = 0
        self.stop_scraping = threading.Event()
        self.scraping_thread = None

        # GUI colors
        self.bg_color = "#2E2E2E"  # Dark Grey Background
        self.fg_color = "#FFFFFF"  # White Text
        self.master.configure(bg=self.bg_color)

        self.setup_logging()
        self.setup_gui()

    def setup_logging(self):
        log_filename = f"search_scraper_log_{datetime.now().strftime('%Y%m%d%H%M%S')}.txt"
        logging.basicConfig(
            filename=log_filename,
            level=logging.DEBUG,
            format="%(asctime)s [%(levelname)s]: %(message)s"
        )
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        logging.getLogger().addHandler(console_handler)

    def setup_gui(self):
        self.create_search_frame()
        self.create_search_operators_text()
        self.create_progress_indicators()
        self.create_buttons()
        self.create_output_format_selection()
        self.create_status_and_log()

    def create_search_frame(self):
        search_frame = tk.Frame(self.master, bg=self.bg_color)
        search_frame.pack(pady=10)

        self.search_query_entry = self._create_labeled_entry(search_frame, "Search Query:", 0)
        self.total_results_per_search_engine_entry = self._create_labeled_entry(search_frame, 
                                                                                "Total Results per Search Engine:", 1)
        
        self.remove_duplicates_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            search_frame,
            text="Remove Duplicates",
            variable=self.remove_duplicates_var,
            bg=self.bg_color,
            fg=self.fg_color,
            selectcolor=self.bg_color,
        ).grid(row=2, column=0, columnspan=2, padx=5, pady=5)

    def _create_labeled_entry(self, parent: tk.Frame, label_text: str, row: int) -> tk.Entry:
        tk.Label(parent, text=label_text, fg=self.fg_color, bg=self.bg_color).grid(row=row, column=0, padx=5, pady=5)
        
        entry = tk.Entry(parent, bg="#3D3D3D", fg=self.fg_color)
        entry.grid(row=row, column=1, padx=5, pady=5)

        setattr(self, f'{label_text.lower().replace(" ", "_").replace(":", "")}_entry', entry)
        
        return entry

    def create_search_operators_text(self):
        tk.Label(self.master, text="Search Operators:", fg=self.fg_color, bg=self.bg_color).pack()
        
        self.search_operators_text = tk.Text(self.master, height=5, bg="#3D3D3D", fg=self.fg_color, wrap=tk.WORD)
        self.search_operators_text.pack()

        for operator, description in SEARCH_OPERATORS.items():
            self.search_operators_text.insert(tk.END, f"{operator} - {description}\n")

    def create_progress_indicators(self):
        tk.Label(self.master, text="Scraping Progress:", fg=self.fg_color, bg=self.bg_color).pack()
        
        self.progress_bar = ttk.Progressbar(self.master, orient="horizontal", length=200, mode="determinate")
        self.progress_bar.pack()
        
        self.progress_percentage_label = tk.Label(self.master, text="Progress: 0%", fg=self.fg_color, bg=self.bg_color)
        self.progress_percentage_label.pack()

    def create_buttons(self):
        self._create_button("Start Scraping", self.start_scraping, "#4CAF50")
        self._create_button("Stop Scraping", self.stop_scraping_command, "#F44336")

    def _create_button(self, text: str, command: callable, bg_color: str):
        tk.Button(
            self.master,
            text=text,
            command=command,
            bg=bg_color,
            fg=self.fg_color,
        ).pack(pady=5)

    def create_output_format_selection(self):
        tk.Label(self.master, text="Output Format:", fg=self.fg_color, bg=self.bg_color).pack()
        
        self.output_format_var = tk.StringVar(value="xlsx")
        
        ttk.Combobox(
            self.master,
            textvariable=self.output_format_var,
            values=["xlsx", "csv"]
        ).pack()

    def create_status_and_log(self):
        self.status_label = tk.Label(self.master, text="", bg=self.bg_color, fg=self.fg_color)
        self.status_label.pack()

        self.log_text = tk.Text(self.master, height=10, bg="#3D3D3D", fg=self.fg_color)
        self.log_text.pack()

    def start_scraping(self):
        query = self.search_query_entry.get().strip()
        
        if not query:
            self.show_error("Please enter a search query.")
            return
        
        try:
            num_results = int(self.total_results_per_search_engine_entry.get())
            if num_results <= 0:
                raise ValueError("Number of results must be a positive integer.")
                
            # Update status and start scraping thread
            self.update_status_label("Scraping in progress...", color="yellow")
            self.stop_scraping.clear()
            
            # Start the scraping thread
            self.scraping_thread = threading.Thread(target=self._scrape_all_engines,
                                                    args=(query, num_results))
            self.scraping_thread.start()

        except ValueError as e:
            self.show_error(str(e))

    def stop_scraping_command(self):
        if self.scraping_thread and self.scraping_thread.is_alive():
            self.stop_scraping.set()
            self.update_status_label("Stopping the scraping process...", color="red")
            
    def _scrape_all_engines(self, query: str, num_results: int):
        try:
            all_results = []
            total_engines = len(SEARCH_ENGINES)
            
            for index, (engine_name, scrape_function_name) in enumerate(SEARCH_ENGINES, 1):
                if self.stop_scraping.is_set():
                    logging.info("Scraping stopped by user.")
                    break

                scrape_function = getattr(self, scrape_function_name)
                engine_results = self._scrape_with_common_logic(
                    engine_name, query.strip(), num_results, scrape_function
                )
                all_results.extend(engine_results)

                self.update_progress(index * num_results, total_engines * num_results)

            if not self.stop_scraping.is_set():
                self._process_results(query, all_results, num_results, total_engines)
            else:
                self.update_status_label("Scraping stopped by user.", color="red")

        except Exception as e:
            self._log_error(f"An error occurred: {str(e)}")
            self.show_error(f"An error occurred: {str(e)}")
            self.update_status_label("Error occurred during scraping", color="red")

        finally:
            self.master.update_idletasks()
            self.master.after(2000, self.clear_status_label)

    def _scrape_with_common_logic(self, engine_name: str, query: str, num_results: int, scrape_function) -> List[Dict]:
        results = []
        try:
            self.update_status_label(f"Scraping {engine_name}...", color="yellow")
            engine_results = scrape_function(query, num_results)
            results.extend(engine_results)
            self.update_status_label(f"{engine_name} scraping complete!", color="green")
        except Exception as e:
            self._log_error(f"Error scraping {engine_name}: {str(e)}")
            self.update_status_label(f"Error scraping {engine_name}", color="red")
        return results

    def _process_results(self, query: str, all_results: List[Dict], num_results: int, num_engines: int):
        total_links_collected = len(all_results)
    
        if self.remove_duplicates_var.get():
            unique_results = self._remove_duplicates(all_results)
        else:
            unique_results = all_results
    
        total_links_removed = total_links_collected - len(unique_results)
    
        self._log_info(f"Total links collected: {total_links_collected}")
        self._log_info(f"Total duplicate links removed: {total_links_removed}")
    
        # Group results by engine
        grouped_results = {}
        for result in unique_results:
            engine = result["Search Engine"]
            if engine not in grouped_results:
                grouped_results[engine] = []
            grouped_results[engine].append(result)
    
        self.total_pages = num_results * num_engines
        self._save_results(query, grouped_results, total_links_collected, total_links_removed)
        self.update_status_label("Scraping and saving complete!", color="green")

    def _remove_duplicates(self, results: List[Dict]) -> List[Dict]:
        """Remove duplicates while maintaining balance between search engines."""
        # Group results by search engine
        engine_results = {}
        for result in results:
            engine = result["Search Engine"]
            if engine not in engine_results:
                engine_results[engine] = []
            engine_results[engine].append(result)

        # Find the minimum number of results across engines
        min_results = min(len(results) for results in engine_results.values())
    
        # Keep track of seen URLs for each engine
        seen_urls = set()
        balanced_results = []

        # Process results from each engine in rotation
        engines = list(engine_results.keys())
        current_index = {engine: 0 for engine in engines}
    
        while True:
            added_any = False
        
            for engine in engines:
                engine_list = engine_results[engine]
                current_idx = current_index[engine]
            
                # Try to add one result from this engine
                while current_idx < len(engine_list):
                    result = engine_list[current_idx]
                    current_idx += 1
                
                    url = result["URL"]
                    if url and url not in seen_urls:
                        seen_urls.add(url)
                        balanced_results.append(result)
                        added_any = True
                        break
            
                current_index[engine] = current_idx
        
            if not added_any:
                break

        return balanced_results

    def _truncate_long_url(self, url: str, max_length=200):
        if len(url) > max_length:
            # Check if it's a Bing redirect URL
            if "bing.com/ck/a" in url:
                # Extract the actual URL from the redirect
                parts = url.split("&u3=")
                if len(parts) > 1:
                    actual_url = parts[1]
                    # Decode the URL if it's encoded
                    actual_url = unquote(actual_url)
                    # Truncate if still too long
                    if len(actual_url) > max_length:
                        truncated_url = actual_url[:max_length] + '...'
                        self._log_warning(f"URL too long. Truncated URL: {truncated_url}")
                        return truncated_url
                    return actual_url
            # For other long URLs, truncate and add an ellipsis
            truncated_url = url[:max_length] + '...'
            self._log_warning(f"URL too long. Truncated URL: {truncated_url}")
            return truncated_url
        return url

    def scrape_bing(self, query: str, num_results: int) -> List[Dict]:
        headers = {"User-Agent": self._get_random_user_agent()}
        bing_results = []
        session = requests.Session()
        offset = 0
        max_pages = 50  # Add a maximum page limit
        current_page = 0
    
        while len(bing_results) < num_results and current_page < max_pages:
            try:
                url = f"https://www.bing.com/search?q={query}&first={offset}"
                response = self._get_response(session, url, headers)
            
                if not response:
                    self._log_warning(f"No response received for offset {offset}")
                    break
                
                soup = BeautifulSoup(response.text, "html.parser")
                search_results = soup.find_all("li", {"class": "b_algo"})
            
                # If no results found on current page, break the loop
                if not search_results:
                    self._log_info(f"No more results found after offset {offset}")
                    break
            
                for result in search_results:
                    if len(bing_results) >= num_results:
                        break
                    extracted_result = self._extract_bing_result(result)
                    if extracted_result:  # Only add if we got a valid result
                        bing_results.append(extracted_result)
            
                offset += 10
                current_page += 1
            
                # Add a small delay between requests
                time.sleep(random.uniform(0.5, 1.5))
            
            except Exception as e:
                self._log_error(f"Error scraping Bing at offset {offset}: {str(e)}")
                break
    
        self._log_info(f"Scraped {len(bing_results)} results from Bing")
        return bing_results

    def _extract_bing_result(self, result) -> Dict:
        title_element = result.find("h2")
        title = title_element.text.strip() if title_element else "No Title"
        
        link_element = result.find("a", href=True)
        link = self._get_final_url(link_element.get("href")) if link_element else None
        
        description_element = result.find("div", {"class": "b_caption"})
        description_element = result.find("div", {"class": "b_caption"})
        description = description_element.text.strip() if description_element else ""
        
        return {
            "Search Engine": "Bing",
            "Title": title,
            "URL": link,
            "Description": description,
            "Page": random.randint(1, 10),
        }

    def scrape_duckduckgo(self, query: str, num_results: int) -> List[Dict]:
        headers = {
            "User-Agent": self._get_random_user_agent(),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Accept-Encoding": "gzip, deflate, br",
            "DNT": "1",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Cache-Control": "max-age=0"
        }
        duckduckgo_results = []
        session = requests.Session()
        offset = 0
        max_pages = 50
        current_page = 0
        max_retries = 3
        retry_delay = 2

        while len(duckduckgo_results) < num_results and current_page < max_pages:
            for attempt in range(max_retries):
                try:
                    if self.stop_scraping.is_set():
                        return duckduckgo_results

                    url = f"https://html.duckduckgo.com/html/?q={query}&s={offset}"
                    response = self._get_response(session, url, headers)
            
                    if not response:
                        self._log_warning(f"No response from DuckDuckGo on attempt {attempt + 1}")
                        time.sleep(retry_delay * (attempt + 1))
                        continue
                
                    if response.status_code == 202:
                        self._log_warning(f"DuckDuckGo returned 202 status, waiting before retry")
                        time.sleep(retry_delay * (attempt + 1))
                        continue

                    soup = BeautifulSoup(response.text, "html.parser")
                    results = soup.find_all("div", class_="result")

                    if not results:
                        self._log_info(f"No results found on page {current_page}")
                        return duckduckgo_results

                    success = False
                    for result in results:
                        if len(duckduckgo_results) >= num_results:
                            return duckduckgo_results
                    
                        extracted_result = self._extract_duckduckgo_result(result, offset)
                        if extracted_result:
                            duckduckgo_results.append(extracted_result)
                            success = True

                    if success:
                        break  # Break retry loop on success
                    else:
                        self._log_warning(f"No valid results extracted on attempt {attempt + 1}")
                        time.sleep(retry_delay * (attempt + 1))

                except Exception as e:
                    self._log_error(f"Error scraping DuckDuckGo: {str(e)}")
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay * (attempt + 1))
                    else:
                        return duckduckgo_results

            offset += 30
            current_page += 1
            time.sleep(random.uniform(1.5, 3))  # Increased delay between pages

        self._log_info(f"Scraped {len(duckduckgo_results)} results from DuckDuckGo")
        return duckduckgo_results

    def _extract_duckduckgo_result(self, result, offset: int) -> Dict:
        try:
            # Find the title and link
            title_element = result.find("h2", class_="result__title")
            link_element = result.find("a", class_="result__a")
        
            if not title_element or not link_element:
                return None
        
            title = title_element.text.strip()
            raw_url = link_element.get('href', '')
        
            # Extract the actual URL from DuckDuckGo's redirect
            if raw_url.startswith('/'):
                # Parse the URL parameters
                parsed = urlparse(raw_url)
                query_params = parse_qs(parsed.query)
                url = query_params.get('uddg', [None])[0]
                if url:
                    url = unquote(url)
            else:
                url = raw_url
        
            # Get description
            description_element = result.find("a", class_="result__snippet")
            description = description_element.text.strip() if description_element else ""
        
            return {
                "Search Engine": "DuckDuckGo",
                "Title": title,
                "URL": url,
                "Description": description,
                "Page": offset // 30 + 1  # Calculate actual page number
            }
        except Exception as e:
            self._log_error(f"Error extracting DuckDuckGo result: {str(e)}")
            return None

    def scrape_yahoo(self, query: str, num_results: int) -> List[Dict]:
        headers = {
            "User-Agent": self._get_random_user_agent(),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Cache-Control": "no-cache"
        }
        yahoo_results = []
        session = requests.Session()
        offset = 1
        max_pages = 50
        current_page = 0

        try:
            # First visit Yahoo homepage to get cookies
            session.get("https://search.yahoo.com", headers=headers, timeout=10)
            time.sleep(2)  # Wait before making search request
        
            while len(yahoo_results) < num_results and current_page < max_pages:
                url = f"https://search.yahoo.com/search?p={query}&b={offset}&pz=10"
                response = session.get(url, headers=headers, timeout=10)
            
                if response.status_code != 200:
                    break
                
                soup = BeautifulSoup(response.text, "html.parser")
                results = soup.find_all("div", {"class": "algo-sr"})  # Updated class name
            
                if not results:
                    results = soup.find_all("div", {"class": "sr"})  # Fallback class
            
                if not results:
                    break
                
                for result in results:
                    if len(yahoo_results) >= num_results:
                        break
                    
                    extracted_result = self._extract_yahoo_result(result)
                    if extracted_result:
                        yahoo_results.append(extracted_result)
                    
                offset += 10
                current_page += 1
                time.sleep(random.uniform(1, 2))
            
        except Exception as e:
            self._log_error(f"Error scraping Yahoo: {str(e)}")
        
        self._log_info(f"Scraped {len(yahoo_results)} results from Yahoo")
        return yahoo_results

    def _extract_yahoo_result(self, result) -> Dict:
        try:
            # Try multiple possible element locations
            title_element = (
                result.find("h3", class_="title") or 
                result.find("h3") or 
                result.find("a", class_="ac-algo")
            )
            title = title_element.text.strip() if title_element else "No Title"
        
            link_element = result.find("a")
            link = None
            if link_element:
                href = link_element.get("href", "")
                # Extract actual URL from Yahoo's redirect
                if "RU=" in href:
                    link = unquote(href.split("RU=")[1].split("/RK=")[0])
                else:
                    link = href
                
            description_element = (
                result.find("div", class_="compText") or 
                result.find("p", class_="lh-16") or
                result.find("p")
            )
            description = description_element.text.strip() if description_element else ""
        
            return {
                "Search Engine": "Yahoo",
                "Title": title,
                "URL": link,
                "Description": description,
                "Page": random.randint(1, 10)
            }
        except Exception as e:
            self._log_error(f"Error extracting Yahoo result: {str(e)}")
            return None

    def scrape_mojeek(self, query: str, num_results: int) -> List[Dict]:
        headers = {"User-Agent": self._get_random_user_agent()}
        session = requests.Session()
        mojeek_results = []
        max_pages = 50  # Add maximum page limit
        current_page = 1

        while len(mojeek_results) < num_results and current_page < max_pages:
            if self.stop_scraping.is_set():
                break

            try:
                url = f"https://www.mojeek.com/search?q={query}&page={current_page}"
                response = self._get_response(session, url, headers)
            
                if not response:
                    self._log_warning(f"No response received for page {current_page}")
                    break

                soup = BeautifulSoup(response.text, "html.parser")
                results = soup.find_all("li", class_=re.compile("r[0-9]+"))

                if not results:
                    self._log_info(f"No more results found after page {current_page}")
                    break

                for result in results:
                    if len(mojeek_results) >= num_results:
                        break
                    
                    extracted_result = self._extract_mojeek_result(result)
                    if extracted_result:
                        mojeek_results.append(extracted_result)

                current_page += 1
                time.sleep(random.uniform(0.5, 1.5))  # Polite delay between requests

            except Exception as e:
                self._log_error(f"Error scraping Mojeek at page {current_page}: {str(e)}")
                break

        self._log_info(f"Scraped {len(mojeek_results)} results from Mojeek")
        return mojeek_results

        
    def _extract_mojeek_result(self, result) -> Dict:
        title_element = result.find("h2")
        title = title_element.text.strip() if title_element else "No Title"
        
        link_element = result.find("a", href=True)
        link = link_element["href"] if link_element else None
        description_element = result.find("p", class_="s")
        description = description_element.text.strip() if description_element else "No Description"

        return {
            "Search Engine": "Mojeek",
            "Title": title,
            "URL": link,
            "Description": description,
            "Page": random.randint(1, 10),  # Random page number
        }

    def _get_final_url(self, url: str) -> str:
        """Extracts the final URL from a potential redirect URL."""
        try:
            parsed_url = urlparse(url)
            query_params = parse_qs(parsed_url.query)

            # Handle Yahoo's specific redirect format
            if "r.search.yahoo.com" in parsed_url.netloc:
                # Extract the RU parameter which contains the actual URL
                if 'RU' in query_params:
                    actual_url = unquote(query_params['RU'][0])
                    return actual_url
        
            # Handle Bing's redirect
            elif parsed_url.netloc == 'bing.com':
                redirect_url_param_names = ['u', 'u3']
                for param_name in redirect_url_param_names:
                    if param_name in query_params:
                        return unquote(query_params[param_name][0])
        
            # Handle DuckDuckGo's redirect
            elif parsed_url.netloc == 'duckduckgo.com':
                if 'uddg' in query_params:
                    return unquote(query_params['uddg'][0])

            return url
        except Exception as e:
            self._log_error(f"Error processing URL {url}: {str(e)}")
            return url


    def _create_session(self):
        return requests.Session()

    def _get_actual_url(self, url: str, session: requests.Session) -> str:
        try:
            response = session.get(url, allow_redirects=True, timeout=10)
            if response.history:
                actual_url = response.url
                self._log_info(f"Redirected URL for Mojeek: {actual_url}")
                return actual_url
            else:
                return url
        except requests.RequestException as e:
            self._log_error(f"Error getting actual URL for Mojeek: {str(e)}")
            return url

    def update_progress(self, current: int, total: int):
        percentage = int((current / total) * 100)
        self.progress_bar["value"] = percentage
        self.progress_percentage_label.config(text=f"Progress: {percentage}%")
        self.master.update_idletasks()

    def _get_response(self, session: requests.Session, url: str, headers: Dict[str, str]) -> requests.Response | None:
        try:
            response = session.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            return response
        except requests.RequestException as e:
            self._log_error(f"Error fetching URL {url}: {str(e)}")
            return None

    def _get_random_user_agent(self) -> str:
        user_agents_file = "User_Agents.txt"
    
        if os.path.exists(user_agents_file):
            with open(user_agents_file, "r") as f:
                user_agents = f.read().splitlines()
        else:
            # Fallback to a default list if the file doesn't exist
            user_agents = [
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15",
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0"
            ]
            self._log_warning(f"User_Agents.txt not found. Using default user agents.")

        return random.choice(user_agents)

    def show_error(self, message: str):
        messagebox.showerror("Error", message)
        self._log_error(message)

    def _log(self, message: str, level: int = logging.INFO):
        logging.log(level, message)
        self._append_to_log(f"[{logging.getLevelName(level)}] {message}")

    def _log_info(self, message: str):
        self._log(message, logging.INFO)

    def _log_warning(self, message: str):
        self._log(message, logging.WARNING)

    def _log_error(self, message: str):
        self._log(message, logging.ERROR)
        self.update_status_label("Error", "red")
        
    def _append_to_log(self, message: str):
        self.log_text.insert(
            tk.END, f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n"
        )
        self.log_text.see(tk.END)

    def update_status_label(self, text: str, color: str = "black"):
        self.status_label.config(text=text, fg=color)
        self._append_to_log(text)

    def clear_status_label(self):
        self.status_label.config(text="")

    def _save_results(self, query: str, results: Dict[str, List[Dict]], total_links_collected: int, total_removed: int):
        try:
            if not any(results.values()):
                self._log_warning("No results to save. Aborting save operation.")
                messagebox.showwarning("No Results", "There are no results to save.")
                return
        
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            output_format = self.output_format_var.get().lower()
            cleaned_query = self._clean_query(query)
            filename = f"{cleaned_query}_results_{timestamp}.{output_format}"
        
            os.makedirs("results", exist_ok=True)
            filename = os.path.join("results", filename)
        
            # Convert results to a flat list for DataFrame
            flat_results = []
            for engine_results in results.values():
                flat_results.extend(engine_results)
            
            df = pd.DataFrame(flat_results)
        
            if output_format == "csv":
                # Reorder columns for CSV
                df = df[['Search Engine', 'Title', 'Page', 'URL', 'Description']]
                df.to_csv(filename, index=False, encoding="utf-8-sig")
            
            elif output_format == "xlsx":
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    # Write main results sheet
                    df = df[['Search Engine', 'Title', 'Page', 'URL', 'Description']]
                    df.to_excel(writer, index=False, sheet_name="Results")
                
                    # Format the Results sheet
                    worksheet = writer.sheets["Results"]
                    for idx, col in enumerate(df.columns):
                        max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                        worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len
                
                    # Make URLs clickable
                    for idx, url in enumerate(df["URL"], start=2):
                        if pd.notna(url):  # Check if URL is not NaN
                            worksheet.cell(row=idx, column=df.columns.get_loc("URL") + 1).hyperlink = url
                
                    # Add summary sheet
                    summary_data = {
                        "Metric": ["Total Links Collected", "Total Duplicate Links Removed", "Total Unique Links"],
                        "Value": [total_links_collected, total_removed, len(df)]
                    }
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, index=False, sheet_name="Summary")
                
                    # Format summary sheet
                    summary_sheet = writer.sheets["Summary"]
                    for idx, col in enumerate(summary_df.columns):
                        max_len = max(summary_df[col].astype(str).map(len).max(), len(col)) + 2
                        summary_sheet.column_dimensions[get_column_letter(idx + 1)].width = max_len
        
            self._log_info(f"File saved successfully to {filename}")
            messagebox.showinfo("Results Saved", 
                        f"Search results saved to {filename}\n"
                        f"Total links collected: {total_links_collected}\n"
                        f"Total duplicate links removed: {total_removed}")
    
        except Exception as e:
            self._log_error(f"Error occurred while saving results: {str(e)}")
            messagebox.showerror("Error", f"An error occurred while saving results: {str(e)}")

    def _save_to_file(self, df: pd.DataFrame, filename: str, output_format: str, total_links_collected: int, total_removed: int):
        try:
            if output_format == "csv":
                # Reorder columns for CSV
                df = df[['Search Engine', 'Title', 'Page', 'URL', 'Description']]
                df.to_csv(filename, index=False, encoding="utf-8-sig")
            elif output_format == "xlsx":
                # Reorder columns for Excel
                df = df[['Search Engine', 'Title', 'Page', 'URL', 'Description']]
                
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    # Write results to Excel
                    df.to_excel(writer, index=False, sheet_name="Results")
                    
                    worksheet = writer.sheets["Results"]
                    for idx, col in enumerate(df.columns):
                        max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                        worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len
                    
                    # Make URLs clickable
                    for idx, url in enumerate(df["URL"], start=2):
                        worksheet.cell(row=idx, column=df.columns.get_loc("URL") + 1).hyperlink = url
                    
                    summary_data = {
                        "Total Links Collected": [total_links_collected],
                        "Total Duplicate Links Removed": [total_removed],
                        "Total Unique Links": [len(df)],
                    }
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, index=False, sheet_name="Summary")
        
            self._log_info(f"File saved successfully to {filename}")
            messagebox.showinfo("Results Saved", f"Search results saved to {filename}\nTotal links collected: {total_links_collected}\nTotal duplicate links removed: {total_removed}")
        except Exception as e:
            self._log_error(f"Error occurred while saving results: {str(e)}")
            messagebox.showerror("Error", f"An error occurred while saving results: {str(e)}")

    @staticmethod
    def _clean_query(query: str) -> str:
        return "".join(c for c in query if c.isalnum() or c.isspace()).replace(" ", "_")

    def apply_dark_theme(self):
        dark_theme = {
            "bg": "#2E2E2E",
            "fg": "#FFFFFF",
            "insertbackground": "#FFFFFF",
        }

        for widget in self.master.winfo_children():
            try:
                widget.config(**dark_theme)
            except tk.TclError:
                pass

        self.progress_bar["style"] = "dark.Horizontal.TProgressbar"
        self.master.tk_setPalette(
            background="#2E2E2E",
            foreground="#FFFFFF",
            activeBackground="#2E2E2E",
            activeForeground="#FFFFFF",
        )

if __name__ == "__main__":
    root = tk.Tk()
    gui = SearchScraperGUI(root)
    gui.apply_dark_theme()
    root.mainloop()
